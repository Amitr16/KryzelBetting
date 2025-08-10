"""
Complete Admin App for GoalServe Sports Betting Platform
Features: Betting Events, User Management, Reports, Report Builder with downloadable reports
"""

from flask import Flask, render_template_string, jsonify, request, send_file, make_response
import sqlite3
import json
from datetime import datetime, timedelta
import os
import io
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# Database path (using the existing GoalServe database)
DATABASE_PATH = 'src/database/app.db'

def get_db_connection():
    """Get database connection"""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def calculate_event_financials(event_id, market_id, sport_name):
    """Calculate max liability and max possible gain for a specific event+market combination"""
    try:
        conn = get_db_connection()
        
        # Get all pending bets for this specific event+market combination
        query = """
        SELECT bet_selection, stake, potential_return, odds
        FROM bets 
        WHERE match_id = ? AND market = ? AND sport_name = ? AND status = 'pending'
        """
        
        bets = conn.execute(query, (event_id, market_id, sport_name)).fetchall()
        conn.close()
        
        if not bets:
            return 0.0, 0.0  # No bets = no liability or gain
        
        # Group bets by selection (outcome)
        selections = {}
        total_stakes = 0
        
        for bet in bets:
            selection = bet['bet_selection']
            stake = float(bet['stake'])
            potential_return = float(bet['potential_return'])
            
            if selection not in selections:
                selections[selection] = {'total_stake': 0, 'total_payout': 0}
            
            selections[selection]['total_stake'] += stake
            selections[selection]['total_payout'] += potential_return
            total_stakes += stake
        
        # Calculate profit/loss for each possible outcome
        outcomes = []
        for selection, data in selections.items():
            # If this selection wins: pay out winners, keep losing stakes
            payout = data['total_payout']
            profit_loss = total_stakes - payout
            outcomes.append(profit_loss)
        
        # Max liability = worst case (most negative outcome)
        max_liability = abs(min(outcomes)) if outcomes else 0.0
        
        # Max possible gain = best case (most positive outcome)  
        max_possible_gain = max(outcomes) if outcomes else 0.0
        
        return max_liability, max_possible_gain
        
    except Exception as e:
        print(f"Error calculating financials: {e}")
        return 0.0, 0.0

def calculate_total_revenue():
    """Calculate total revenue from settled bets"""
    try:
        conn = get_db_connection()
        
        # Calculate revenue from settled bets
        # Revenue = Total stakes from losing bets - Total payouts to winning bets
        query = """
        SELECT 
            SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) as total_stakes_lost,
            SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as total_payouts_won
        FROM bets 
        WHERE status IN ('won', 'lost')
        """
        
        result = conn.execute(query).fetchone()
        conn.close()
        
        total_stakes_lost = result['total_stakes_lost'] or 0
        total_payouts_won = result['total_payouts_won'] or 0
        
        # Revenue = Money kept from losing bets - Extra money paid to winners
        total_revenue = total_stakes_lost - total_payouts_won
        
        return total_revenue
        
    except Exception as e:
        print(f"Error calculating total revenue: {e}")
        return 0.0

def generate_report_data(report_type, date_from=None, date_to=None, sport_filter=None):
    """Generate report data based on type and filters"""
    try:
        conn = get_db_connection()
        
        # Build date filter
        date_filter = ""
        params = []
        
        if date_from:
            date_filter += " AND DATE(created_at) >= ?"
            params.append(date_from)
        if date_to:
            date_filter += " AND DATE(created_at) <= ?"
            params.append(date_to)
        if sport_filter:
            date_filter += " AND sport_name = ?"
            params.append(sport_filter)
        
        if report_type == 'betting_summary':
            query = f"""
            SELECT 
                DATE(created_at) as bet_date,
                sport_name,
                COUNT(*) as total_bets,
                SUM(stake) as total_stakes,
                COUNT(CASE WHEN status = 'won' THEN 1 END) as won_bets,
                COUNT(CASE WHEN status = 'lost' THEN 1 END) as lost_bets,
                COUNT(CASE WHEN status = 'pending' THEN 1 END) as pending_bets,
                SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) - 
                SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as revenue
            FROM bets
            WHERE 1=1 {date_filter}
            GROUP BY DATE(created_at), sport_name
            ORDER BY bet_date DESC, sport_name
            """
            
        elif report_type == 'user_activity':
            query = f"""
            SELECT 
                u.username,
                u.email,
                u.balance,
                COUNT(b.id) as total_bets,
                COALESCE(SUM(CASE WHEN b.status IN ('won', 'lost') THEN b.stake ELSE 0 END), 0) as total_staked,
                COUNT(CASE WHEN b.status = 'won' THEN 1 END) as won_bets,
                COUNT(CASE WHEN b.status = 'lost' THEN 1 END) as lost_bets,
                COALESCE(SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END), 0) as payout,
                COALESCE(SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END), 0) - COALESCE(SUM(CASE WHEN b.status IN ('won', 'lost') THEN b.stake ELSE 0 END), 0) as user_profit,
                u.created_at as joined_date
            FROM users u
            LEFT JOIN bets b ON u.id = b.user_id
            WHERE 1=1 {date_filter.replace('created_at', 'b.created_at')}
            GROUP BY u.id, u.username, u.email, u.balance, u.created_at
            ORDER BY total_bets DESC
            """
            
        elif report_type == 'financial_overview':
            query = f"""
            SELECT 
                DATE(created_at) as report_date,
                COUNT(*) as daily_bets,
                SUM(stake) as daily_stakes,
                SUM(CASE WHEN status = 'won' THEN potential_return ELSE 0 END) as daily_payouts,
                SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) as daily_revenue_from_losses,
                SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) - 
                SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as daily_net_revenue
            FROM bets
            WHERE 1=1 {date_filter}
            GROUP BY DATE(created_at)
            ORDER BY report_date DESC
            """
            
        elif report_type == 'sport_performance':
            query = f"""
            SELECT 
                sport_name,
                COUNT(*) as total_bets,
                SUM(stake) as total_stakes,
                COUNT(CASE WHEN status = 'won' THEN 1 END) as won_bets,
                COUNT(CASE WHEN status = 'lost' THEN 1 END) as lost_bets,
                COUNT(CASE WHEN status = 'pending' THEN 1 END) as pending_bets,
                SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) - 
                SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as sport_revenue,
                ROUND(COUNT(CASE WHEN status = 'won' THEN 1 END) * 100.0 / COUNT(*), 2) as win_rate
            FROM bets
            WHERE 1=1 {date_filter}
            GROUP BY sport_name
            ORDER BY sport_revenue DESC
            """
        
        data = conn.execute(query, params).fetchall()
        conn.close()
        
        # Convert to list of dictionaries
        return [dict(row) for row in data]
        
    except Exception as e:
        print(f"Error generating report data: {e}")
        return []

def create_pdf_report(data, report_type, title):
    """Create PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 12))
    
    # Generated date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1
    )
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
    story.append(Spacer(1, 20))
    
    if not data:
        story.append(Paragraph("No data available for the selected criteria.", styles['Normal']))
    else:
        # Create table
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data:
            table_data.append([str(row[key]) for key in headers])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def create_excel_report(data, report_type, title):
    """Create Excel report"""
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = report_type.replace('_', ' ').title()
    
    # Title
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=16, bold=True)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    # Generated date
    worksheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    worksheet['A2'].font = Font(size=10)
    worksheet['A2'].alignment = Alignment(horizontal='center')
    
    if not data:
        worksheet['A4'] = "No data available for the selected criteria."
    else:
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, header in enumerate(headers, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=row_data[header])
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

def create_csv_report(data, report_type, title):
    """Create CSV report"""
    buffer = io.StringIO()
    
    if not data:
        buffer.write("No data available for the selected criteria.\n")
    else:
        writer = csv.DictWriter(buffer, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    # Convert to bytes buffer
    bytes_buffer = io.BytesIO()
    bytes_buffer.write(buffer.getvalue().encode('utf-8'))
    bytes_buffer.seek(0)
    return bytes_buffer

@app.route('/')
def admin_dashboard():
    """Main admin dashboard"""
    return render_template_string(ADMIN_TEMPLATE)

@app.route('/api/betting-events')
def get_betting_events():
    """Get all available betting events from JSON files with improved parsing"""
    try:
        import os
        import json
        
        # Get database connection for checking disabled events
        conn = get_db_connection()
        
        # Path to Sports Pre Match directory
        sports_dir = 'Sports Pre Match'
        
        all_events = []
        
        # Get pagination parameters
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # Get filter and search parameters
        sport_filter = request.args.get('sport', '')
        market_filter = request.args.get('market', '')
        search_query = request.args.get('search', '').lower()
        sort_by = request.args.get('sort_by', 'event_id')
        sort_order = request.args.get('sort_order', 'asc')
        
        # Iterate through all sport directories
        for sport_name in os.listdir(sports_dir):
            sport_path = os.path.join(sports_dir, sport_name)
            if not os.path.isdir(sport_path):
                continue
                
            # Skip if sport filter doesn't match
            if sport_filter and sport_filter.lower() != sport_name.lower():
                continue
            
            # Look for JSON file
            json_path = os.path.join(sport_path, f"{sport_name}_odds.json")
            if not os.path.exists(json_path):
                continue
                
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                # Extract sport display name from metadata
                sport_display_name = sport_name.title()
                if 'metadata' in data and 'display_name' in data['metadata']:
                    sport_display_name = data['metadata']['display_name']
                
                # Navigate to matches
                if 'odds_data' in data and 'scores' in data['odds_data']:
                    scores = data['odds_data']['scores']
                    if 'categories' in scores:
                        for category in scores['categories']:
                            if 'matches' in category:
                                category_name = category.get('name', 'Unknown Category')
                                
                                for match in category['matches']:
                                    event_id = match.get('id', '')
                                    
                                    # Extract team names properly
                                    localteam = match.get('localteam', {})
                                    awayteam = match.get('awayteam', {})
                                    
                                    home_team = localteam.get('name', 'Unknown') if isinstance(localteam, dict) else str(localteam)
                                    away_team = awayteam.get('name', 'Unknown') if isinstance(awayteam, dict) else str(awayteam)
                                    
                                    event_name = f"{home_team} vs {away_team}"
                                    
                                    # Skip if search query doesn't match
                                    if search_query and search_query not in event_name.lower():
                                        continue
                                    
                                    # Process odds/markets
                                    if 'odds' in match:
                                        for odds_entry in match['odds']:
                                            market_name = odds_entry.get('value', 'Unknown Market')
                                            market_id = odds_entry.get('id', 'unknown')
                                            
                                            # Skip if market filter doesn't match
                                            if market_filter and market_filter.lower() != market_name.lower():
                                                continue
                                                
                                            # Create betting event entry
                                            betting_event = {
                                                'id': f"{event_id}_{market_id}",
                                                'event_id': f"{event_id}_{market_id}",  # Use EventID_MarketID format
                                                'sport': sport_display_name,
                                                'event_name': event_name,
                                                'market': market_name,
                                                'market_display': market_name,
                                                'category': category_name,
                                                'odds_data': odds_entry,
                                                'is_active': True,  # Will be updated below
                                                'date': match.get('date', ''),
                                                'time': match.get('time', ''),
                                                'status': match.get('status', 'Unknown')
                                            }
                                            
                                            # Check if this event is disabled
                                            event_key = f"{event_id}_{market_id}"
                                            disabled_check = conn.execute(
                                                'SELECT * FROM disabled_events WHERE event_key = ?', 
                                                (event_key,)
                                            ).fetchone()
                                            betting_event['is_active'] = disabled_check is None
                                            
                                            # Calculate financial metrics for this event/market
                                            max_liability, max_possible_gain = calculate_event_financials(event_id, market_id, sport_name)
                                            betting_event['max_liability'] = max_liability
                                            betting_event['max_possible_gain'] = max_possible_gain
                                            
                                            all_events.append(betting_event)
                                        
            except Exception as e:
                print(f"Error reading {json_path}: {e}")
                continue
        
        # Sort events
        reverse = sort_order == 'desc'
        if sort_by == 'event_id':
            all_events.sort(key=lambda x: int(x['event_id']) if x['event_id'].isdigit() else 0, reverse=reverse)
        elif sort_by == 'sport':
            all_events.sort(key=lambda x: x['sport'], reverse=reverse)
        elif sort_by == 'event_name':
            all_events.sort(key=lambda x: x['event_name'], reverse=reverse)
        elif sort_by == 'market':
            all_events.sort(key=lambda x: x['market'], reverse=reverse)
        elif sort_by == 'max_liability':
            all_events.sort(key=lambda x: x['max_liability'], reverse=reverse)
        elif sort_by == 'max_possible_gain':
            all_events.sort(key=lambda x: x['max_possible_gain'], reverse=reverse)
        elif sort_by == 'status':
            all_events.sort(key=lambda x: x['is_active'], reverse=reverse)
        
        # Apply pagination
        total_events = len(all_events)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_events = all_events[start_idx:end_idx]
        
        # Get unique sports and markets for filters
        unique_sports = list(set(event['sport'] for event in all_events))
        unique_markets = list(set(event['market'] for event in all_events))
        
        # Calculate summary statistics
        active_events = len([e for e in all_events if e['is_active']])
        
        # Calculate liability per unique event+market combination
        event_market_liabilities = {}
        for event in all_events:
            # Use the full event_id (EventID_MarketID format) as the unique key
            full_event_id = event['event_id']
            
            if event['max_liability'] > 0:
                event_market_liabilities[full_event_id] = event['max_liability']
        
        total_liability = sum(event_market_liabilities.values())
        total_revenue = calculate_total_revenue()  # Calculate from settled bets
        
        conn.close()
        
        return jsonify({
            'events': paginated_events,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total_events,
                'pages': (total_events + per_page - 1) // per_page
            },
            'summary': {
                'total_events': total_events,
                'active_events': active_events,
                'total_liability': total_liability,
                'total_revenue': total_revenue,  # Changed from total_gain
                'unique_sports': len(unique_sports),
                'unique_markets': len(unique_markets)
            },
            'filters': {
                'sports': sorted(unique_sports),
                'markets': sorted(unique_markets)
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/betting-events/<event_key>/toggle', methods=['POST'])
def toggle_event_status(event_key):
    """Toggle event active status"""
    try:
        conn = get_db_connection()
        
        # Check if event is currently disabled
        disabled_event = conn.execute(
            'SELECT * FROM disabled_events WHERE event_key = ?', 
            (event_key,)
        ).fetchone()
        
        if disabled_event:
            # Event is disabled, enable it by removing from disabled_events
            conn.execute('DELETE FROM disabled_events WHERE event_key = ?', (event_key,))
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'Event enabled successfully',
                'is_active': True
            })
        else:
            # Event is active, disable it by adding to disabled_events
            # Parse event_key to extract components (format: event_id_market)
            parts = event_key.split('_')
            if len(parts) >= 2:
                event_id = '_'.join(parts[:-1])
                market = parts[-1]
            else:
                event_id = event_key
                market = 'unknown'
            
            conn.execute(
                'INSERT OR REPLACE INTO disabled_events (event_key, sport, event_name, market) VALUES (?, ?, ?, ?)',
                (event_key, 'unknown', 'Unknown Event', market)
            )
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'Event disabled successfully',
                'is_active': False
            })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/users')
def get_users():
    """Get all users with betting statistics"""
    try:
        conn = get_db_connection()
        
        # Get sorting parameters
        sort_by = request.args.get('sort_by', 'id')
        sort_order = request.args.get('sort_order', 'asc')
        
        # Validate sort_by parameter
        valid_sort_fields = {
            'id': 'u.id',
            'username': 'u.username', 
            'email': 'u.email',
            'balance': 'u.balance',
            'total_bets': 'total_bets',
            'total_staked': 'total_staked',
            'payout': 'payout',
            'cumulative_profit': 'cumulative_profit',
            'joined': 'u.created_at',
            'status': 'u.is_active'
        }
        
        sort_field = valid_sort_fields.get(sort_by, 'u.id')
        sort_direction = 'ASC' if sort_order.lower() == 'asc' else 'DESC'
        
        # Get users with betting stats
        query = f"""
        SELECT 
            u.id,
            u.username,
            u.email,
            u.balance,
            u.created_at,
            u.is_active,
            COUNT(b.id) as total_bets,
            COALESCE(SUM(CASE WHEN b.status IN ('won', 'lost') THEN b.stake ELSE 0 END), 0) as total_staked,
            COALESCE(SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END), 0) as payout,
            COALESCE(SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END), 0) - COALESCE(SUM(CASE WHEN b.status IN ('won', 'lost') THEN b.stake ELSE 0 END), 0) as cumulative_profit
        FROM users u
        LEFT JOIN bets b ON u.id = b.user_id
        GROUP BY u.id, u.username, u.email, u.balance, u.created_at, u.is_active
        ORDER BY {sort_field} {sort_direction}
        """
        
        users = conn.execute(query).fetchall()
        conn.close()
        
        users_list = []
        for user in users:
            users_list.append({
                'id': user['id'],
                'username': user['username'],
                'email': user['email'],
                'balance': f"${user['balance']:.2f}",
                'total_bets': user['total_bets'],
                'total_staked': f"${user['total_staked']:.2f}",
                'payout': f"${user['payout']:.2f}",
                'cumulative_profit': f"${user['cumulative_profit']:.2f}",
                'joined': user['created_at'][:10] if user['created_at'] else 'Unknown',
                'is_active': user['is_active']
            })
        
        return jsonify({'users': users_list})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>/toggle', methods=['POST'])
def toggle_user_status(user_id):
    """Toggle user active status"""
    try:
        conn = get_db_connection()
        
        # Get current user status
        user = conn.execute('SELECT is_active FROM users WHERE id = ?', (user_id,)).fetchone()
        
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        # Toggle status
        new_status = not user['is_active']
        conn.execute('UPDATE users SET is_active = ? WHERE id = ?', (new_status, user_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'User {"enabled" if new_status else "disabled"} successfully',
            'is_active': new_status
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/overview')
def get_reports_overview():
    """Get comprehensive reports overview"""
    try:
        conn = get_db_connection()
        
        # Get date range for reports (last 30 days)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)
        
        # Total bets and revenue
        total_query = """
        SELECT 
            COUNT(*) as total_bets,
            SUM(stake) as total_stakes,
            SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as total_payouts,
            SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) as total_revenue_from_losses,
            COUNT(CASE WHEN status = 'pending' THEN 1 END) as pending_bets,
            COUNT(CASE WHEN status = 'won' THEN 1 END) as won_bets,
            COUNT(CASE WHEN status = 'lost' THEN 1 END) as lost_bets
        FROM bets
        """
        
        totals = conn.execute(total_query).fetchone()
        
        # Daily revenue for the last 30 days
        daily_query = """
        SELECT 
            DATE(created_at) as bet_date,
            COUNT(*) as daily_bets,
            SUM(stake) as daily_stakes,
            SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) - 
            SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as daily_revenue
        FROM bets
        WHERE created_at >= date('now', '-30 days')
        GROUP BY DATE(created_at)
        ORDER BY bet_date DESC
        """
        
        daily_data = conn.execute(daily_query).fetchall()
        
        # Sport-wise performance
        sport_query = """
        SELECT 
            sport_name,
            COUNT(*) as bets_count,
            SUM(stake) as total_stakes,
            SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) - 
            SUM(CASE WHEN status = 'won' THEN potential_return - stake ELSE 0 END) as sport_revenue
        FROM bets
        GROUP BY sport_name
        ORDER BY sport_revenue DESC
        """
        
        sport_data = conn.execute(sport_query).fetchall()
        
        # Top users by activity
        user_query = """
        SELECT 
            u.username,
            u.email,
            COUNT(b.id) as total_bets,
            SUM(b.stake) as total_staked,
            SUM(CASE WHEN b.status = 'won' THEN b.potential_return - b.stake ELSE 0 END) as user_profit
        FROM users u
        LEFT JOIN bets b ON u.id = b.user_id
        GROUP BY u.id, u.username, u.email
        HAVING COUNT(b.id) > 0
        ORDER BY total_bets DESC
        LIMIT 10
        """
        
        top_users = conn.execute(user_query).fetchall()
        
        conn.close()
        
        # Calculate total revenue (same as calculate_total_revenue function)
        total_revenue = (totals['total_revenue_from_losses'] or 0) - (totals['total_payouts'] or 0)
        
        return jsonify({
            'overview': {
                'total_bets': totals['total_bets'] or 0,
                'total_stakes': totals['total_stakes'] or 0,
                'total_revenue': total_revenue,
                'pending_bets': totals['pending_bets'] or 0,
                'won_bets': totals['won_bets'] or 0,
                'lost_bets': totals['lost_bets'] or 0,
                'win_rate': (totals['won_bets'] / max(totals['total_bets'], 1)) * 100 if totals['total_bets'] else 0
            },
            'daily_data': [
                {
                    'date': row['bet_date'],
                    'bets': row['daily_bets'],
                    'stakes': row['daily_stakes'] or 0,
                    'revenue': row['daily_revenue'] or 0
                }
                for row in daily_data
            ],
            'sport_performance': [
                {
                    'sport': row['sport_name'],
                    'bets': row['bets_count'],
                    'stakes': row['total_stakes'] or 0,
                    'revenue': row['sport_revenue'] or 0
                }
                for row in sport_data
            ],
            'top_users': [
                {
                    'username': row['username'],
                    'email': row['email'],
                    'total_bets': row['total_bets'],
                    'total_staked': row['total_staked'] or 0,
                    'profit': row['user_profit'] or 0
                }
                for row in top_users
            ]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reports/available-sports')
def get_available_sports():
    """Get list of available sports for report filtering"""
    try:
        conn = get_db_connection()
        sports = conn.execute('SELECT DISTINCT sport_name FROM bets ORDER BY sport_name').fetchall()
        conn.close()
        
        return jsonify({
            'sports': [row['sport_name'] for row in sports if row['sport_name']]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reports/generate', methods=['POST'])
def generate_report():
    """Generate and download report"""
    try:
        data = request.get_json()
        
        report_type = data.get('report_type')
        format_type = data.get('format', 'pdf')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        sport_filter = data.get('sport_filter')
        
        # Generate report data
        report_data = generate_report_data(report_type, date_from, date_to, sport_filter)
        
        # Create title
        title_map = {
            'betting_summary': 'Betting Summary Report',
            'user_activity': 'User Activity Report',
            'financial_overview': 'Financial Overview Report',
            'sport_performance': 'Sport Performance Report'
        }
        title = title_map.get(report_type, 'Custom Report')
        
        if date_from or date_to:
            date_range = f" ({date_from or 'Start'} to {date_to or 'End'})"
            title += date_range
        
        # Generate report based on format
        if format_type == 'pdf':
            buffer = create_pdf_report(report_data, report_type, title)
            mimetype = 'application/pdf'
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
        elif format_type == 'excel':
            buffer = create_excel_report(report_data, report_type, title)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        elif format_type == 'csv':
            buffer = create_csv_report(report_data, report_type, title)
            mimetype = 'text/csv'
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        else:
            return jsonify({'error': 'Invalid format type'}), 400
        
        return send_file(
            buffer,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# HTML Template with Report Builder
ADMIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GoalServe Admin Dashboard</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }
        
        .header {
            background: rgba(0, 0, 0, 0.8);
            color: white;
            padding: 1rem 2rem;
            display: flex;
            align-items: center;
            gap: 1rem;
        }
        
        .header h1 {
            font-size: 1.5rem;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            padding: 2rem;
        }
        
        .nav-tabs {
            display: flex;
            gap: 1rem;
            margin-bottom: 2rem;
        }
        
        .nav-tab {
            padding: 0.75rem 1.5rem;
            background: rgba(255, 255, 255, 0.9);
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            transition: all 0.3s ease;
        }
        
        .nav-tab.active {
            background: #4CAF50;
            color: white;
        }
        
        .nav-tab:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
        }
        
        .content-section {
            display: none;
            background: rgba(255, 255, 255, 0.95);
            border-radius: 12px;
            padding: 2rem;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        }
        
        .content-section.active {
            display: block;
        }
        
        .summary-cards {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
            margin-bottom: 2rem;
        }
        
        .summary-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 1.5rem;
            border-radius: 8px;
            text-align: center;
        }
        
        .summary-card h3 {
            font-size: 0.9rem;
            margin-bottom: 0.5rem;
            opacity: 0.9;
        }
        
        .summary-card .value {
            font-size: 1.8rem;
            font-weight: bold;
        }
        
        .controls {
            display: flex;
            gap: 1rem;
            margin-bottom: 2rem;
            flex-wrap: wrap;
            align-items: center;
        }
        
        .controls select, .controls input, .controls button {
            padding: 0.5rem;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 0.9rem;
        }
        
        .controls button {
            background: #4CAF50;
            color: white;
            border: none;
            cursor: pointer;
            font-weight: 600;
        }
        
        .controls button:hover {
            background: #45a049;
        }
        
        .table-container {
            overflow-x: auto;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            background: white;
        }
        
        th, td {
            padding: 0.75rem;
            text-align: left;
            border-bottom: 1px solid #eee;
        }
        
        th {
            background: #f8f9fa;
            font-weight: 600;
            cursor: pointer;
            user-select: none;
            position: relative;
        }
        
        th:hover {
            background: #e9ecef;
        }
        
        th.sortable::after {
            content: ' ↕';
            opacity: 0.5;
        }
        
        th.sort-asc::after {
            content: ' ↑';
            opacity: 1;
        }
        
        th.sort-desc::after {
            content: ' ↓';
            opacity: 1;
        }
        
        .status-badge {
            padding: 0.25rem 0.5rem;
            border-radius: 12px;
            font-size: 0.8rem;
            font-weight: 600;
        }
        
        .status-active {
            background: #d4edda;
            color: #155724;
        }
        
        .status-disabled {
            background: #f8d7da;
            color: #721c24;
        }
        
        .liability {
            color: #dc3545;
            font-weight: 600;
        }
        
        .gain {
            color: #28a745;
            font-weight: 600;
        }
        
        .revenue {
            color: #007bff;
            font-weight: 600;
        }
        
        .action-btn {
            padding: 0.25rem 0.75rem;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.8rem;
            font-weight: 600;
        }
        
        .btn-enable {
            background: #28a745;
            color: white;
        }
        
        .btn-disable {
            background: #dc3545;
            color: white;
        }
        
        .btn-download {
            background: #17a2b8;
            color: white;
            margin-left: 0.5rem;
        }
        
        .pagination {
            display: flex;
            justify-content: center;
            align-items: center;
            gap: 1rem;
            margin-top: 2rem;
        }
        
        .pagination button {
            padding: 0.5rem 1rem;
            border: 1px solid #ddd;
            background: white;
            cursor: pointer;
            border-radius: 4px;
        }
        
        .pagination button:hover:not(:disabled) {
            background: #f8f9fa;
        }
        
        .pagination button:disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }
        
        .pagination .current-page {
            font-weight: 600;
            color: #667eea;
        }
        
        .loading {
            text-align: center;
            padding: 2rem;
            color: #666;
        }
        
        .error {
            background: #f8d7da;
            color: #721c24;
            padding: 1rem;
            border-radius: 4px;
            margin-bottom: 1rem;
        }
        
        .report-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 2rem;
            margin-bottom: 2rem;
        }
        
        .report-section {
            background: white;
            border-radius: 8px;
            padding: 1.5rem;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }
        
        .report-section h3 {
            margin-bottom: 1rem;
            color: #333;
            border-bottom: 2px solid #667eea;
            padding-bottom: 0.5rem;
        }
        
        .metric-row {
            display: flex;
            justify-content: space-between;
            padding: 0.5rem 0;
            border-bottom: 1px solid #eee;
        }
        
        .metric-row:last-child {
            border-bottom: none;
        }
        
        .metric-label {
            font-weight: 500;
            color: #666;
        }
        
        .metric-value {
            font-weight: 600;
            color: #333;
        }
        
        .report-builder {
            background: white;
            border-radius: 8px;
            padding: 2rem;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
            margin-bottom: 2rem;
        }
        
        .form-group {
            margin-bottom: 1rem;
        }
        
        .form-group label {
            display: block;
            margin-bottom: 0.5rem;
            font-weight: 600;
            color: #333;
        }
        
        .form-row {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
            margin-bottom: 1rem;
        }
        
        .download-buttons {
            display: flex;
            gap: 1rem;
            margin-top: 1rem;
        }
        
        .download-buttons button {
            padding: 0.75rem 1.5rem;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
            transition: all 0.3s ease;
        }
        
        .btn-pdf {
            background: #dc3545;
            color: white;
        }
        
        .btn-excel {
            background: #28a745;
            color: white;
        }
        
        .btn-csv {
            background: #ffc107;
            color: #212529;
        }
        
        .download-buttons button:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>🏆 GoalServe Admin Dashboard</h1>
    </div>
    
    <div class="container">
        <div class="nav-tabs">
            <button class="nav-tab active" onclick="showSection('betting-events')">📊 Betting Events</button>
            <button class="nav-tab" onclick="showSection('user-management')">👥 User Management</button>
            <button class="nav-tab" onclick="showSection('reports')">📈 Reports</button>
            <button class="nav-tab" onclick="showSection('report-builder')">🔧 Report Builder</button>
        </div>
        
        <!-- Betting Events Section -->
        <div id="betting-events" class="content-section active">
            <h2>Betting Events Management</h2>
            
            <div class="summary-cards">
                <div class="summary-card">
                    <h3>Total Events</h3>
                    <div class="value" id="total-events">-</div>
                </div>
                <div class="summary-card">
                    <h3>Active Events</h3>
                    <div class="value" id="active-events">-</div>
                </div>
                <div class="summary-card">
                    <h3>Total Liability</h3>
                    <div class="value" id="max-liability">$0.00</div>
                </div>
                <div class="summary-card">
                    <h3>Total Revenue</h3>
                    <div class="value revenue" id="total-revenue">$0.00</div>
                </div>
            </div>
            
            <div class="controls">
                <select id="sport-filter">
                    <option value="">All Sports</option>
                </select>
                <select id="market-filter">
                    <option value="">All Markets</option>
                </select>
                <input type="text" id="search-input" placeholder="Search event names...">
                <select id="per-page-select">
                    <option value="10">10 per page</option>
                    <option value="20" selected>20 per page</option>
                    <option value="50">50 per page</option>
                    <option value="100">100 per page</option>
                </select>
                <button onclick="loadBettingEvents()">🔄 Refresh Events</button>
            </div>
            
            <div class="table-container">
            <table>
                <thead>
                    <tr>
                            <th class="sortable" data-sort="event_id">Event ID</th>
                            <th class="sortable" data-sort="sport">Sport</th>
                            <th class="sortable" data-sort="event_name">Event Name</th>
                            <th class="sortable" data-sort="market">Market</th>
                            <th class="sortable" data-sort="max_liability">Max Liability</th>
                            <th class="sortable" data-sort="max_possible_gain">Max Possible Gain</th>
                            <th class="sortable" data-sort="status">Status</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                    <tbody id="events-table-body">
                    <tr><td colspan="8" class="loading">Loading events...</td></tr>
                </tbody>
            </table>
        </div>
        
            <div class="pagination" id="events-pagination">
                <!-- Pagination will be inserted here -->
            </div>
        </div>
        
        <!-- User Management Section -->
        <div id="user-management" class="content-section">
            <h2>User Management</h2>
            
            <div class="controls">
                <button onclick="loadUsers()">🔄 Refresh Users</button>
            </div>
            
            <div class="table-container">
            <table>
                <thead>
                    <tr>
                            <th class="sortable" data-sort="id">ID</th>
                            <th class="sortable" data-sort="username">Username</th>
                            <th class="sortable" data-sort="email">Email</th>
                            <th class="sortable" data-sort="balance">Balance</th>
                            <th class="sortable" data-sort="total_bets">Total Bets</th>
                            <th class="sortable" data-sort="total_staked">Total Staked</th>
                            <th class="sortable" data-sort="payout">Payout</th>
                            <th class="sortable" data-sort="cumulative_profit">Cumulative Profit</th>
                            <th class="sortable" data-sort="joined">Joined</th>
                            <th class="sortable" data-sort="status">Status</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                    <tbody id="users-table-body">
                        <tr><td colspan="11" class="loading">Loading users...</td></tr>
                </tbody>
            </table>
            </div>
        </div>
        
        <!-- Reports Section -->
        <div id="reports" class="content-section">
            <h2>Reports & Analytics</h2>
            
            <div class="summary-cards">
                <div class="summary-card">
                    <h3>Total Bets</h3>
                    <div class="value" id="report-total-bets">-</div>
                </div>
                <div class="summary-card">
                    <h3>Total Stakes</h3>
                    <div class="value" id="report-total-stakes">$0.00</div>
                </div>
                <div class="summary-card">
                    <h3>Total Revenue</h3>
                    <div class="value revenue" id="report-total-revenue">$0.00</div>
                </div>
                <div class="summary-card">
                    <h3>Win Rate</h3>
                    <div class="value" id="report-win-rate">0%</div>
                </div>
            </div>
            
            <div class="controls">
                <button onclick="loadReports()">🔄 Refresh Reports</button>
            </div>
            
            <div class="report-grid">
                <div class="report-section">
                    <h3>📊 Betting Overview</h3>
                    <div id="betting-overview">
                        <div class="loading">Loading overview...</div>
                    </div>
                </div>
                
                <div class="report-section">
                    <h3>🏆 Sport Performance</h3>
                    <div id="sport-performance">
                        <div class="loading">Loading sport data...</div>
                    </div>
                </div>
                
                <div class="report-section">
                    <h3>👑 Top Users</h3>
                    <div id="top-users">
                        <div class="loading">Loading user data...</div>
                    </div>
                </div>
                
                <div class="report-section">
                    <h3>📈 Recent Activity</h3>
                    <div id="recent-activity">
                        <div class="loading">Loading activity...</div>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Report Builder Section -->
        <div id="report-builder" class="content-section">
            <h2>🔧 Report Builder</h2>
            
            <div class="report-builder">
                <h3>Generate Custom Reports</h3>
                
                <div class="form-group">
                    <label for="report-type">Report Type:</label>
                    <select id="report-type">
                        <option value="betting_summary">Betting Summary</option>
                        <option value="user_activity">User Activity</option>
                        <option value="financial_overview">Financial Overview</option>
                        <option value="sport_performance">Sport Performance</option>
                    </select>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label for="date-from">Date From:</label>
                        <input type="date" id="date-from">
                    </div>
                    <div class="form-group">
                        <label for="date-to">Date To:</label>
                        <input type="date" id="date-to">
                    </div>
                    <div class="form-group">
                        <label for="sport-filter-report">Sport Filter:</label>
                        <select id="sport-filter-report">
                            <option value="">All Sports</option>
                        </select>
                    </div>
                </div>
                
                <div class="download-buttons">
                    <button class="btn-pdf" onclick="downloadReport('pdf')">📄 Download PDF</button>
                    <button class="btn-excel" onclick="downloadReport('excel')">📊 Download Excel</button>
                    <button class="btn-csv" onclick="downloadReport('csv')">📋 Download CSV</button>
                </div>
            </div>
            
            <div class="report-section">
                <h3>📋 Report Descriptions</h3>
                <div class="metric-row">
                    <span class="metric-label">Betting Summary:</span>
                    <span class="metric-value">Daily betting activity by sport</span>
                </div>
                <div class="metric-row">
                    <span class="metric-label">User Activity:</span>
                    <span class="metric-value">User statistics and betting behavior</span>
                </div>
                <div class="metric-row">
                    <span class="metric-label">Financial Overview:</span>
                    <span class="metric-value">Daily revenue and financial metrics</span>
                </div>
                <div class="metric-row">
                    <span class="metric-label">Sport Performance:</span>
                    <span class="metric-value">Revenue and performance by sport</span>
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let currentPage = 1;
        let currentSort = 'event_id';
        let currentSortOrder = 'asc';
        
        // User sorting variables
        let userCurrentSort = 'id';
        let userCurrentSortOrder = 'asc';
        
        function showSection(sectionId) {
            // Hide all sections
            document.querySelectorAll('.content-section').forEach(section => {
                section.classList.remove('active');
            });
            
            // Remove active class from all tabs
            document.querySelectorAll('.nav-tab').forEach(tab => {
                tab.classList.remove('active');
            });
            
            // Show selected section
            document.getElementById(sectionId).classList.add('active');
            
            // Add active class to clicked tab
            event.target.classList.add('active');
            
            // Load data for the section
            if (sectionId === 'betting-events') {
                loadBettingEvents();
            } else if (sectionId === 'user-management') {
                loadUsers();
            } else if (sectionId === 'reports') {
                loadReports();
            } else if (sectionId === 'report-builder') {
                loadReportBuilder();
            }
        }
        
        function loadBettingEvents(page = 1) {
            console.log('Loading betting events, page:', page);
            currentPage = page;
            
            const sportFilter = document.getElementById('sport-filter').value;
            const marketFilter = document.getElementById('market-filter').value;
            const searchQuery = document.getElementById('search-input').value;
            const perPage = document.getElementById('per-page-select').value;
            
            const params = new URLSearchParams({
                page: page,
                per_page: perPage,
                sort_by: currentSort,
                sort_order: currentSortOrder
            });
            
                if (sportFilter) params.append('sport', sportFilter);
                if (marketFilter) params.append('market', marketFilter);
                if (searchQuery) params.append('search', searchQuery);
                
            fetch(`/api/betting-events?${params}`)
                .then(response => response.json())
                .then(data => {
                    if (data.error) {
                        document.getElementById('events-table-body').innerHTML = 
                            `<tr><td colspan="8" class="error">Error: ${data.error}</td></tr>`;
                        return;
                    }
                    
                    // Update summary cards
                    document.getElementById('total-events').textContent = data.summary.total_events;
                    document.getElementById('active-events').textContent = data.summary.active_events;
                    document.getElementById('max-liability').textContent = `$${data.summary.total_liability.toFixed(2)}`;
                    document.getElementById('total-revenue').textContent = `$${data.summary.total_revenue.toFixed(2)}`;
                    
                    // Update filters
                    updateFilters(data.filters);
                    
                    // Update table
                    updateEventsTable(data.events);
                    
                    // Update pagination
                    updatePagination(data.pagination, 'events');
                })
                .catch(error => {
                    console.error('Error loading events:', error);
                    document.getElementById('events-table-body').innerHTML = 
                        `<tr><td colspan="8" class="error">Failed to load events</td></tr>`;
                });
        }
        
        function updateFilters(filters) {
            const sportFilter = document.getElementById('sport-filter');
            const marketFilter = document.getElementById('market-filter');
            
            // Update sport filter
            const currentSport = sportFilter.value;
            sportFilter.innerHTML = '<option value="">All Sports</option>';
            filters.sports.forEach(sport => {
                            const option = document.createElement('option');
                option.value = sport.toLowerCase();
                option.textContent = sport;
                if (sport.toLowerCase() === currentSport) option.selected = true;
                sportFilter.appendChild(option);
            });
            
            // Update market filter
            const currentMarket = marketFilter.value;
            marketFilter.innerHTML = '<option value="">All Markets</option>';
            filters.markets.forEach(market => {
                            const option = document.createElement('option');
                option.value = market.toLowerCase();
                option.textContent = market;
                if (market.toLowerCase() === currentMarket) option.selected = true;
                marketFilter.appendChild(option);
            });
        }
        
        function updateEventsTable(events) {
            const tbody = document.getElementById('events-table-body');
            
            if (events.length === 0) {
                tbody.innerHTML = '<tr><td colspan="8" class="loading">No events found</td></tr>';
                return;
            }
            
            tbody.innerHTML = events.map(event => `
                <tr>
                            <td>${event.event_id}</td>
                            <td>${event.sport}</td>
                            <td>${event.event_name}</td>
                    <td>${event.market}</td>
                    <td class="liability">$${event.max_liability.toFixed(2)}</td>
                    <td class="gain">$${event.max_possible_gain.toFixed(2)}</td>
                    <td>
                        <span class="status-badge ${event.is_active ? 'status-active' : 'status-disabled'}">
                            ${event.is_active ? 'Active' : 'Disabled'}
                        </span>
                    </td>
                    <td>
                        <button class="action-btn ${event.is_active ? 'btn-disable' : 'btn-enable'}" 
                                onclick="toggleEvent('${event.id}')">
                                    ${event.is_active ? 'Disable' : 'Enable'}
                                </button>
                            </td>
                </tr>
            `).join('');
        }
        
        function updatePagination(pagination, type) {
            const container = document.getElementById(`${type}-pagination`);
            
            const prevDisabled = pagination.page <= 1 ? 'disabled' : '';
            const nextDisabled = pagination.page >= pagination.pages ? 'disabled' : '';
            
            container.innerHTML = `
                <button ${prevDisabled} onclick="loadBettingEvents(${pagination.page - 1})">Previous</button>
                <span class="current-page">Page ${pagination.page} of ${pagination.pages}</span>
                <span>(${pagination.total} total events)</span>
                <button ${nextDisabled} onclick="loadBettingEvents(${pagination.page + 1})">Next</button>
            `;
        }
        
        function toggleEvent(eventId) {
            fetch(`/api/betting-events/${eventId}/toggle`, {
                method: 'POST'
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    loadBettingEvents(currentPage);
                } else {
                    alert('Error: ' + data.error);
                }
            })
            .catch(error => {
                console.error('Error toggling event:', error);
                alert('Failed to toggle event status');
            });
        }
        
        function loadUsers() {
            console.log('Loading users...');
            const params = new URLSearchParams({
                sort_by: userCurrentSort,
                sort_order: userCurrentSortOrder
            });
            
            console.log('Fetching from:', `/api/users?${params}`);
            fetch(`/api/users?${params}`)
                .then(response => {
                    console.log('Response status:', response.status);
                    return response.json();
                })
                .then(data => {
                    console.log('Users data:', data);
                    if (data.error) {
                        document.getElementById('users-table-body').innerHTML = 
                            `<tr><td colspan="11" class="error">Error: ${data.error}</td></tr>`;
                        return;
                    }
                    
                    updateUsersTable(data.users);
                })
                .catch(error => {
                    console.error('Error loading users:', error);
                    document.getElementById('users-table-body').innerHTML = 
                        `<tr><td colspan="11" class="error">Failed to load users</td></tr>`;
                });
        }
        
        function updateUsersTable(users) {
            const tbody = document.getElementById('users-table-body');
            
            if (users.length === 0) {
                tbody.innerHTML = '<tr><td colspan="11" class="loading">No users found</td></tr>';
                return;
            }
            
            tbody.innerHTML = users.map(user => `
                <tr>
                            <td>${user.id}</td>
                            <td>${user.username}</td>
                            <td>${user.email}</td>
                    <td>${user.balance}</td>
                            <td>${user.total_bets}</td>
                    <td>${user.total_staked}</td>
                    <td class="gain">${user.payout}</td>
                    <td class="${user.cumulative_profit.startsWith('-') ? 'liability' : 'gain'}">${user.cumulative_profit}</td>
                    <td>${user.joined}</td>
                    <td>
                        <span class="status-badge ${user.is_active ? 'status-active' : 'status-disabled'}">
                            ${user.is_active ? 'Active' : 'Disabled'}
                        </span>
                    </td>
                    <td>
                        <button class="action-btn ${user.is_active ? 'btn-disable' : 'btn-enable'}" 
                                onclick="toggleUser(${user.id})">
                                    ${user.is_active ? 'Disable' : 'Enable'}
                                </button>
                            </td>
                </tr>
            `).join('');
        }
        
        function toggleUser(userId) {
            fetch(`/api/users/${userId}/toggle`, {
                method: 'POST'
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    loadUsers();
                } else {
                    alert('Error: ' + data.error);
                }
            })
            .catch(error => {
                console.error('Error toggling user:', error);
                alert('Failed to toggle user status');
            });
        }
        
        function loadReports() {
            fetch('/api/reports/overview')
                .then(response => response.json())
                .then(data => {
                    if (data.error) {
                        console.error('Error loading reports:', data.error);
                        return;
                    }
                    
                    // Update summary cards
                    document.getElementById('report-total-bets').textContent = data.overview.total_bets;
                    document.getElementById('report-total-stakes').textContent = `$${data.overview.total_stakes.toFixed(2)}`;
                    document.getElementById('report-total-revenue').textContent = `$${data.overview.total_revenue.toFixed(2)}`;
                    document.getElementById('report-win-rate').textContent = `${data.overview.win_rate.toFixed(1)}%`;
                    
                    // Update betting overview
                    const overviewHtml = `
                        <div class="metric-row">
                            <span class="metric-label">Pending Bets:</span>
                            <span class="metric-value">${data.overview.pending_bets}</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Won Bets:</span>
                            <span class="metric-value gain">${data.overview.won_bets}</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Lost Bets:</span>
                            <span class="metric-value liability">${data.overview.lost_bets}</span>
                        </div>
                    `;
                    document.getElementById('betting-overview').innerHTML = overviewHtml;
                    
                    // Update sport performance
                    const sportHtml = data.sport_performance.slice(0, 5).map(sport => `
                        <div class="metric-row">
                            <span class="metric-label">${sport.sport}:</span>
                            <span class="metric-value revenue">$${sport.revenue.toFixed(2)}</span>
                        </div>
                    `).join('');
                    document.getElementById('sport-performance').innerHTML = sportHtml || '<div class="loading">No sport data</div>';
                    
                    // Update top users
                    const usersHtml = data.top_users.slice(0, 5).map(user => `
                        <div class="metric-row">
                            <span class="metric-label">${user.username}:</span>
                            <span class="metric-value">${user.total_bets} bets</span>
                        </div>
                    `).join('');
                    document.getElementById('top-users').innerHTML = usersHtml || '<div class="loading">No user data</div>';
                    
                    // Update recent activity
                    const activityHtml = data.daily_data.slice(0, 5).map(day => `
                        <div class="metric-row">
                            <span class="metric-label">${day.date}:</span>
                            <span class="metric-value revenue">$${day.revenue.toFixed(2)}</span>
                        </div>
                    `).join('');
                    document.getElementById('recent-activity').innerHTML = activityHtml || '<div class="loading">No activity data</div>';
                })
                .catch(error => {
                    console.error('Error loading reports:', error);
                });
        }
        
        function loadReportBuilder() {
            // Load available sports for report filtering
            fetch('/api/reports/available-sports')
                .then(response => response.json())
                .then(data => {
                    const sportFilter = document.getElementById('sport-filter-report');
                    sportFilter.innerHTML = '<option value="">All Sports</option>';
                    
                    if (data.sports) {
                        data.sports.forEach(sport => {
                            const option = document.createElement('option');
                            option.value = sport;
                            option.textContent = sport;
                            sportFilter.appendChild(option);
                        });
                    }
                })
                .catch(error => {
                    console.error('Error loading sports:', error);
                });
        }
        
        function downloadReport(format) {
            const reportType = document.getElementById('report-type').value;
            const dateFrom = document.getElementById('date-from').value;
            const dateTo = document.getElementById('date-to').value;
            const sportFilter = document.getElementById('sport-filter-report').value;
            
            const requestData = {
                report_type: reportType,
                format: format,
                date_from: dateFrom || null,
                date_to: dateTo || null,
                sport_filter: sportFilter || null
            };
            
            fetch('/api/reports/generate', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json'
                },
                body: JSON.stringify(requestData)
            })
            .then(response => {
                if (response.ok) {
                    return response.blob();
                } else {
                    throw new Error('Failed to generate report');
                }
            })
            .then(blob => {
                // Create download link
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.style.display = 'none';
                a.href = url;
                
                // Set filename based on format
                const timestamp = new Date().toISOString().slice(0, 19).replace(/:/g, '-');
                const extension = format === 'excel' ? 'xlsx' : format;
                a.download = `${reportType}_${timestamp}.${extension}`;
                
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                document.body.removeChild(a);
            })
            .catch(error => {
                console.error('Error downloading report:', error);
                alert('Failed to download report. Please try again.');
            });
        }
        
        // Add sorting functionality
        document.addEventListener('DOMContentLoaded', function() {
            // Betting events sorting
            document.querySelectorAll('#betting-events th.sortable').forEach(th => {
                th.addEventListener('click', function() {
                    const sortBy = this.dataset.sort;
                    
                    if (currentSort === sortBy) {
                        currentSortOrder = currentSortOrder === 'asc' ? 'desc' : 'asc';
                    } else {
                        currentSort = sortBy;
                        currentSortOrder = 'asc';
                    }
                    
                    // Update visual indicators
                    document.querySelectorAll('#betting-events th.sortable').forEach(header => {
                        header.classList.remove('sort-asc', 'sort-desc');
                    });
                    
                    this.classList.add(currentSortOrder === 'asc' ? 'sort-asc' : 'sort-desc');
                    
                    loadBettingEvents(1);
                });
            });
            
            // User management sorting
            document.querySelectorAll('#user-management th.sortable').forEach(th => {
                th.addEventListener('click', function() {
                    const sortBy = this.dataset.sort;
                    
                    if (userCurrentSort === sortBy) {
                        userCurrentSortOrder = userCurrentSortOrder === 'asc' ? 'desc' : 'asc';
                    } else {
                        userCurrentSort = sortBy;
                        userCurrentSortOrder = 'asc';
                    }
                    
                    // Update visual indicators
                    document.querySelectorAll('#user-management th.sortable').forEach(header => {
                        header.classList.remove('sort-asc', 'sort-desc');
                    });
                    
                    this.classList.add(userCurrentSortOrder === 'asc' ? 'sort-asc' : 'sort-desc');
                    
                    loadUsers();
                });
            });
            
            // Add event listeners for filters
            document.getElementById('sport-filter').addEventListener('change', () => loadBettingEvents(1));
            document.getElementById('market-filter').addEventListener('change', () => loadBettingEvents(1));
            document.getElementById('search-input').addEventListener('input', () => loadBettingEvents(1));
            document.getElementById('per-page-select').addEventListener('change', () => loadBettingEvents(1));
        
        // Load initial data
        loadBettingEvents();
        });
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)

